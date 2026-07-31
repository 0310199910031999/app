import datetime
import hashlib
import re
import secrets
import shutil
import zipfile
from pathlib import Path
from typing import Optional

from config import settings
from mainContext.application.use_cases.Formats.generate_fobc01_pdf_use_case import GenerateFoBc01PdfUseCase
from mainContext.application.dtos.export_dto import (
    ExportDocumentRowDTO,
    ExportJobCompleteDTO,
    ExportJobDTO,
    ExportJobProgressDTO,
)
from mainContext.infrastructure.adapters.AppUserRepo import AppUserRepoImpl
from mainContext.infrastructure.adapters.ExportDocumentCollectorRepo import ExportDocumentCollectorRepoImpl
from mainContext.infrastructure.adapters.ExportJobRepo import ExportJobRepoImpl
from mainContext.infrastructure.adapters.Formats.fo_bc_01_repo import FOBC01RepoImpl
from mainContext.infrastructure.adapters.Formats.fo_cr_02_repo import FOCR02RepoImpl
from mainContext.infrastructure.adapters.Formats.fo_em_01_repo import FOEM01RepoImpl
from mainContext.infrastructure.adapters.Formats.fo_im_01_repo import FOIM01RepoImpl
from mainContext.infrastructure.adapters.Formats.fo_im_03_repo import FOIM03RepoImpl
from mainContext.infrastructure.adapters.Formats.fo_le_01_repo import FOLE01RepoImpl
from mainContext.infrastructure.adapters.Formats.fo_os_01_repo import FOOS01RepoImpl
from mainContext.infrastructure.adapters.Formats.fo_pc_02_repo import FOPC02RepoImpl
from mainContext.infrastructure.adapters.Formats.fo_pp_02_repo import FOPP02RepoImpl
from mainContext.infrastructure.adapters.Formats.fo_sc_01_repo import FOSC01RepoImpl
from mainContext.infrastructure.adapters.Formats.fo_sp_01_repo import FOSP01RepoImpl
from mainContext.infrastructure.adapters.weasyprint_pdf_adapter import WeasyPrintPdfAdapter
from shared.email_service import EmailService


class ExportBatchService:
    MONTH_NAMES = {
        1: 'Enero',
        2: 'Febrero',
        3: 'Marzo',
        4: 'Abril',
        5: 'Mayo',
        6: 'Junio',
        7: 'Julio',
        8: 'Agosto',
        9: 'Septiembre',
        10: 'Octubre',
        11: 'Noviembre',
        12: 'Diciembre',
    }
    STATUS_MESSAGES = {
        'queued': 'La exportación está en cola.',
        'collecting': 'Recolectando documentos para la exportación.',
        'rendering_pdfs': 'Generando archivos PDF.',
        'building_excel': 'Construyendo el Excel consolidado.',
        'compressing': 'Comprimiendo el lote en ZIP.',
        'notifying': 'Enviando la notificación por correo.',
        'completed': 'La exportación está lista.',
        'expired': 'El enlace de descarga ha expirado.',
        'failed': 'La exportación falló.',
    }

    def __init__(
        self,
        job_repo: ExportJobRepoImpl,
        collector: ExportDocumentCollectorRepoImpl,
        app_user_repo: AppUserRepoImpl,
    ):
        self.job_repo = job_repo
        self.collector = collector
        self.app_user_repo = app_user_repo
        self.pdf_generator = WeasyPrintPdfAdapter()
        self._renderer_registry = {
            'fo_bc_01': self._render_fobc01_pdf,
            'fo_cr_02': (FOCR02RepoImpl, 'get_focr02_by_id', self.pdf_generator.generate_focr02_pdf),
            'fo_em_01': (FOEM01RepoImpl, 'get_foem01_by_id', self.pdf_generator.generate_foem01_pdf),
            'fo_im_01': (FOIM01RepoImpl, 'get_foim01_by_id', self.pdf_generator.generate_foim01_pdf),
            'fo_im_03': (FOIM03RepoImpl, 'get_foim03_by_id', self.pdf_generator.generate_foim03_pdf),
            'fo_le_01': (FOLE01RepoImpl, 'get_fole01_by_id', self.pdf_generator.generate_fole01_pdf),
            'fo_os_01': (FOOS01RepoImpl, 'get_foos01_by_id', self.pdf_generator.generate_foos01_pdf),
            'fo_pc_02': (FOPC02RepoImpl, 'get_fopc02_by_id', self.pdf_generator.generate_fopc02_pdf),
            'fo_pp_02': (FOPP02RepoImpl, 'get_fopp02_by_id', self.pdf_generator.generate_fopp02_pdf),
            'fo_sc_01': (FOSC01RepoImpl, 'get_fosc01_by_id', self.pdf_generator.generate_fosc01_pdf),
            'fo_sp_01': (FOSP01RepoImpl, 'get_fosp01_by_id', self.pdf_generator.generate_fosp01_pdf),
        }

    @classmethod
    def status_message(cls, stage: str, status: str, error_message: Optional[str] = None) -> str:
        if status == 'failed' and error_message:
            return error_message
        return cls.STATUS_MESSAGES.get(stage) or cls.STATUS_MESSAGES.get(status) or 'Estado de exportación actualizado.'

    @staticmethod
    def cleanup_job_artifacts(zip_path: Optional[str]):
        if not zip_path:
            return

        zip_file = Path(zip_path)
        if zip_file.exists():
            job_dir = zip_file.parent.parent
            shutil.rmtree(job_dir, ignore_errors=True)

    def process_job(self, job_id: str):
        job = self.job_repo.get_job_by_id(job_id)
        if not job:
            raise ValueError(f'No se encontró el export job con id {job_id}')

        started_at = datetime.datetime.now()
        self.job_repo.update_job_progress(
            job_id,
            ExportJobProgressDTO(
                status='processing',
                stage='collecting',
                progress_pct=0,
                processed_documents=0,
                total_documents=0,
                started_at=started_at,
            ),
        )

        work_dir = None
        try:
            documents = self.collector.collect_documents(job)
            total_documents = len(documents)

            self.job_repo.update_job_progress(
                job_id,
                ExportJobProgressDTO(
                    status='processing',
                    stage='rendering_pdfs',
                    progress_pct=5 if total_documents else 20,
                    processed_documents=0,
                    total_documents=total_documents,
                    started_at=started_at,
                ),
            )

            work_dir, bundle_dir = self._prepare_job_dirs(job_id)
            excel_rows = []

            for index, document in enumerate(documents, start=1):
                pdf_bytes = self._render_pdf(document)
                pdf_path = self._pdf_output_path(work_dir, document)
                pdf_path.parent.mkdir(parents=True, exist_ok=True)
                pdf_path.write_bytes(pdf_bytes)
                excel_rows.append(document.excel_row)

                progress_pct = 10 + int((index / max(total_documents, 1)) * 70)
                self.job_repo.update_job_progress(
                    job_id,
                    ExportJobProgressDTO(
                        status='processing',
                        stage='rendering_pdfs',
                        progress_pct=min(progress_pct, 80),
                        processed_documents=index,
                        total_documents=total_documents,
                        started_at=started_at,
                    ),
                )

            self.job_repo.update_job_progress(
                job_id,
                ExportJobProgressDTO(
                    status='processing',
                    stage='building_excel',
                    progress_pct=85,
                    processed_documents=total_documents,
                    total_documents=total_documents,
                    started_at=started_at,
                ),
            )

            client_name = self.collector.get_client_name(job.client_id)
            excel_filename = f'Reporte de Servicios - {self._sanitize_filename(client_name)}.xlsx'
            excel_path = work_dir / excel_filename
            self._build_excel(excel_rows, excel_path, client_name)

            self.job_repo.update_job_progress(
                job_id,
                ExportJobProgressDTO(
                    status='processing',
                    stage='compressing',
                    progress_pct=95,
                    processed_documents=total_documents,
                    total_documents=total_documents,
                    started_at=started_at,
                ),
            )

            zip_filename = f'export_{job_id}.zip'
            zip_path = bundle_dir / zip_filename
            self._build_zip(work_dir, zip_path)

            raw_token = secrets.token_urlsafe(32)
            token_hash = self._hash_token(raw_token)
            expires_at = datetime.datetime.now() + datetime.timedelta(minutes=settings.EXPORT_URL_TTL_MINUTES)

            self.job_repo.update_job_progress(
                job_id,
                ExportJobProgressDTO(
                    status='processing',
                    stage='notifying',
                    progress_pct=98,
                    processed_documents=total_documents,
                    total_documents=total_documents,
                    started_at=started_at,
                ),
            )

            self.job_repo.complete_job(
                job_id,
                ExportJobCompleteDTO(
                    zip_filename=zip_filename,
                    zip_path=str(zip_path.resolve()),
                    zip_size_bytes=zip_path.stat().st_size,
                    download_token_hash=token_hash,
                    token_expires_at=expires_at,
                    processed_documents=total_documents,
                    total_documents=total_documents,
                ),
            )

            email_sent = self._send_completion_email(job, raw_token, expires_at)
            if not email_sent:
                self.job_repo.update_job_progress(
                    job_id,
                    ExportJobProgressDTO(
                        status='completed',
                        stage='completed',
                        progress_pct=100,
                        processed_documents=total_documents,
                        total_documents=total_documents,
                        error_message='No se pudo enviar el correo de exportación.',
                    ),
                )

            shutil.rmtree(work_dir, ignore_errors=True)
        except Exception as e:
            if work_dir is not None:
                shutil.rmtree(work_dir.parent, ignore_errors=True)
            self.job_repo.fail_job(job_id, str(e))
            raise

    def _prepare_job_dirs(self, job_id: str):
        root_dir = Path(settings.EXPORT_TMP_DIR).resolve() / job_id
        work_dir = root_dir / 'work'
        bundle_dir = root_dir / 'bundle'
        work_dir.mkdir(parents=True, exist_ok=True)
        bundle_dir.mkdir(parents=True, exist_ok=True)
        return work_dir, bundle_dir

    def _sanitize_path_segment(self, value: str) -> str:
        sanitized = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', '_', value or '').strip(' .')
        return sanitized or 'sin_nombre'

    def _sanitize_filename(self, value: str) -> str:
        sanitized = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', '_', value or '').strip(' .')
        return sanitized or 'sin_nombre'

    def _pdf_output_path(self, work_dir: Path, document: ExportDocumentRowDTO) -> Path:
        month_name = self.MONTH_NAMES[document.date_created.month]
        return (
            work_dir
            / self._sanitize_path_segment(document.folder_equipment_name)
            / document.date_created.strftime('%Y')
            / self._sanitize_path_segment(month_name)
            / self._sanitize_path_segment(document.format_folder_name)
            / self._sanitize_path_segment(document.filename)
        )

    def _render_fobc01_pdf(self, document_id: int) -> bytes:
        repo = FOBC01RepoImpl(self.collector.db)
        use_case = GenerateFoBc01PdfUseCase(self.pdf_generator, repo)
        return use_case.execute(fobc01_id=document_id)

    def _render_pdf(self, document: ExportDocumentRowDTO) -> bytes:
        if document.format_key not in self._renderer_registry:
            raise ValueError(f'Formato no soportado para PDF: {document.format_key}')

        renderer_entry = self._renderer_registry[document.format_key]
        if callable(renderer_entry):
            return renderer_entry(document.document_id)

        repo_cls, detail_method, pdf_method = renderer_entry
        repo = repo_cls(self.collector.db)
        detail = getattr(repo, detail_method)(document.document_id)
        if not detail:
            raise ValueError(f'No se encontró el detalle del documento {document.document_id} para {document.format_label}')
        return pdf_method(detail.__dict__)

    def _build_excel(self, rows, output_path: Path, client_name: str = ''):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise RuntimeError('openpyxl es requerido para generar el Excel consolidado') from exc

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Export'

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        data_font = Font(name='Arial', size=10)
        watermark_font = Font(name='Arial', size=8, italic=True, color='A0A0A0')
        title_font = Font(name='Arial', size=16, bold=True, color='1F4E79')
        subtitle_font = Font(name='Arial', size=12, bold=True, color='333333')
        date_font = Font(name='Arial', size=10, color='666666')

        headers = [
            'ID',
            'Equipo',
            'Fecha',
            'Tipo de servicio / Nombre de Formato',
            'Servicios realizados',
            'Desperfectos',
            'Técnico / Empleado',
            'Nombre de Recepción del Servicio',
        ]

        ws_title_row = 1
        ws_client_row = 2
        ws_date_row = 3
        ws_watermark_row = 4
        ws_header_row = 5
        total_cols = len(headers)

        last_col_letter = get_column_letter(total_cols)
        merge_range = f'A{ws_title_row}:{last_col_letter}{ws_title_row}'
        worksheet.merge_cells(merge_range)
        title_cell = worksheet.cell(row=ws_title_row, column=1, value='Reporte de Servicios')
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        worksheet.row_dimensions[ws_title_row].height = 30

        merge_range = f'A{ws_client_row}:{last_col_letter}{ws_client_row}'
        worksheet.merge_cells(merge_range)
        client_cell = worksheet.cell(row=ws_client_row, column=1, value=f'Cliente: {client_name}')
        client_cell.font = subtitle_font
        client_cell.alignment = Alignment(horizontal='left', vertical='center')

        merge_range = f'A{ws_date_row}:{last_col_letter}{ws_date_row}'
        worksheet.merge_cells(merge_range)
        now_str = datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        date_cell = worksheet.cell(row=ws_date_row, column=1, value=f'Fecha de generación: {now_str}')
        date_cell.font = date_font
        date_cell.alignment = Alignment(horizontal='left', vertical='center')

        merge_range = f'A{ws_watermark_row}:{last_col_letter}{ws_watermark_row}'
        worksheet.merge_cells(merge_range)
        watermark_cell = worksheet.cell(
            row=ws_watermark_row,
            column=1,
            value='Generado mediante los servidores DAL Dealer - Thanks for deal with us',
        )
        watermark_cell.font = watermark_font
        watermark_cell.alignment = Alignment(horizontal='center', vertical='center')
        worksheet.row_dimensions[ws_watermark_row].height = 18

        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=ws_header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        worksheet.row_dimensions[ws_header_row].height = 28

        for row_idx, row_data in enumerate(rows, start=ws_header_row + 1):
            max_lines = 1
            for col_idx, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell_val = str(cell.value or '')
                lines = cell_val.count('\n') + 1
                if lines > max_lines:
                    max_lines = lines
            worksheet.row_dimensions[row_idx].height = max(15, max_lines * 15)

        for col_idx, header in enumerate(headers, start=1):
            max_length = len(header)
            for row_idx in range(ws_header_row + 1, ws_header_row + 1 + len(rows)):
                cell_val = worksheet.cell(row=row_idx, column=col_idx).value
                max_length = max(max_length, len(str(cell_val or '')))
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

        if rows:
            last_data_row = ws_header_row + len(rows)
            worksheet.auto_filter.ref = f'A{ws_header_row}:{last_col_letter}{last_data_row}'

        worksheet.sheet_properties.tabColor = '1F4E79'

        workbook.save(output_path)

    def _build_zip(self, source_dir: Path, zip_path: Path):
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as archive:
            for file_path in source_dir.rglob('*'):
                if file_path.is_file():
                    archive.write(file_path, file_path.relative_to(source_dir))

    def _hash_token(self, raw_token: str) -> str:
        return hashlib.sha256(raw_token.encode('utf-8')).hexdigest()

    def _send_completion_email(self, job: ExportJobDTO, raw_token: str, expires_at: datetime.datetime) -> bool:
        app_user = self.app_user_repo.get_app_user_by_id(job.requested_by_user_id)
        if not app_user or not app_user.email:
            return False

        export_base_url = (settings.EXPORT_BASE_URL or settings.BASE_URL).rstrip('/')
        download_url = f"{export_base_url}/exports/download/{raw_token}"
        subject = f'Exportación lista {job.id}'
        message = f"""
        <html>
        <body style=\"font-family: Arial, sans-serif;\">
            <div style=\"max-width: 600px; margin: 0 auto; padding: 20px;\">
                <h3>Tu exportación está lista</h3>
                <p>Se generó correctamente el archivo ZIP solicitado.</p>
                <p><strong>Job:</strong> {job.id}</p>
                <p><strong>Expira:</strong> {expires_at.strftime('%d/%m/%Y %H:%M:%S')}</p>
                <p>
                    <a href=\"{download_url}\" style=\"display: inline-block; padding: 10px 16px; background: #0066cc; color: #fff; text-decoration: none; border-radius: 4px;\">
                        Descargar ZIP
                    </a>
                </p>
            </div>
        </body>
        </html>
        """
        return EmailService.send_email(
            to=app_user.email,
            subject=subject,
            message=message,
            company_id=90,
            html=True,
        )
